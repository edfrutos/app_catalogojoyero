import os
import certifi
import secrets
from datetime import datetime, timedelta
import tempfile
import zipfile

from flask import (
    Flask, request, render_template, redirect, url_for,
    send_from_directory, session, flash, send_file
)
import openpyxl
from openpyxl import load_workbook, Workbook
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from pymongo import MongoClient
from flask_mail import Mail, Message
from bson import ObjectId

# Forzar el uso del bundle de certificados de certifi
os.environ['SSL_CERT_FILE'] = certifi.where()

# -------------------------------------------
# CONFIGURACIÓN FLASK
# -------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("MBZl1W45ute3UEMCXPlL9JzcR7XsTeUi-4ZI6KCd79M", "CAMBIA_ESTA_CLAVE_EN_PRODUCCION")

# Carpeta para imágenes del catálogo
app.config["UPLOAD_FOLDER"] = os.path.join(app.root_path, "imagenes_subidas")
if not os.path.exists(app.config["UPLOAD_FOLDER"]):
    os.makedirs(app.config["UPLOAD_FOLDER"])

# Carpeta para hojas de cálculo (tablas)
SPREADSHEET_FOLDER = os.path.join(app.root_path, "spreadsheets")
if not os.path.exists(SPREADSHEET_FOLDER):
    os.makedirs(SPREADSHEET_FOLDER)

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# -------------------------------------------
# CONFIGURACIÓN DE EMAIL (Flask-Mail)
# -------------------------------------------
app.config["MAIL_SERVER"] = "smtp-relay.brevo.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = "admin@edefrutos.me"
app.config["MAIL_PASSWORD"] = "Rmp3UXwsIkvA0c1d"
app.config["MAIL_DEFAULT_SENDER"] = ("Administrador", "admin@edefrutos.me")
app.config["MAIL_DEBUG"] = True
mail = Mail(app)

# -------------------------------------------
# CONEXIÓN A MONGODB ATLAS
# -------------------------------------------
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi

MONGO_URI = "mongodb+srv://edfrutos:rYjwUC6pUNrLtbaI@cluster0.pmokh.mongodb.net/?retryWrites=true&w=majority"

client = MongoClient(
    MONGO_URI,
    tls=True,
    tlsCAFile=certifi.where(),
    server_api=ServerApi('1')
)

try:
    client.admin.command('ping')
    print("Pinged your deployment. You successfully connected to MongoDB!")
except Exception as e:
    print(e)

db = client["app_catalogojoyero"]
users_collection = db["users"]
resets_collection = db["password_resets"]
spreadsheets_collection = db["spreadsheets"]

# -------------------------------------------
# FUNCIONES AUXILIARES PARA HOJAS DE CÁLCULO
# -------------------------------------------
def leer_datos_excel(filename):
    if not os.path.exists(filename):
        return []

    wb = load_workbook(filename, read_only=True)
    hoja = wb.active
    data = []

    headers = [cell.value for cell in hoja[1]]
    for row in hoja.iter_rows(min_row=2, values_only=True):
        registro = {headers[i]: row[i] for i in range(len(headers))}
        data.append(registro)

    wb.close()
    return data


def escribir_datos_excel(data, filename):
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Datos"

    # Obtener encabezados desde la sesión
    headers = session.get("selected_headers", ["Número", "Descripción", "Peso", "Valor", "Imagen1", "Imagen2", "Imagen3"])

    # Asegurar que "Número" siempre esté presente
    if "Número" not in headers:
        headers.insert(0, "Número")

    # Escribir los encabezados en la hoja de cálculo
    for col_idx, header in enumerate(headers, start=1):
        hoja.cell(row=1, column=col_idx, value=header)

    # Escribir los datos en la hoja
    for fila, item in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            valor = item.get(header, "")

            # Si el valor es una lista (imágenes), convertir a string
            if isinstance(valor, list):
                valor = ", ".join([v for v in valor if v])

            hoja.cell(row=fila, column=col_idx, value=valor)

    wb.save(filename)
    wb.close()

def get_current_spreadsheet():
    filename = session.get("selected_table")
    if not filename:
        return None
    return os.path.join(SPREADSHEET_FOLDER, filename)

# -------------------------------------------
# NUEVA RUTA: PÁGINA DE BIENVENIDA (MANUAL DE USO)
# -------------------------------------------
@app.route("/welcome")
def welcome():
    # Si el usuario no ha iniciado sesión, redirige al login
    if "usuario" not in session:
        return redirect(url_for("login"))
    return render_template("welcome.html")

# -------------------------------------------
# RUTAS DE AUTENTICACIÓN
# -------------------------------------------
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        nombre = request.form.get("nombre").strip()
        email = request.form.get("email").strip().lower()
        password = request.form.get("password").strip()
        if users_collection.find_one({"email": email}):
            return "Error: Ese email ya está registrado. <a href='/register'>Volver</a>"
        hashed = generate_password_hash(password)
        nuevo_usuario = {"nombre": nombre, "email": email, "password": hashed}
        users_collection.insert_one(nuevo_usuario)
        return "Registro exitoso. <a href='/login'>Iniciar Sesión</a>"
    else:
        return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        login_input = request.form.get("login_input").strip()
        password = request.form.get("password").strip()
        usuario = users_collection.find_one({
            "$or": [
                {"nombre": {"$regex": f"^{login_input}$", "$options": "i"}},
                {"email": {"$regex": f"^{login_input}$", "$options": "i"}}
            ]
        })
        if not usuario:
            return "Error: Usuario no encontrado. <a href='/login'>Reintentar</a>"
        if check_password_hash(usuario["password"], password):
            session["usuario"] = usuario["nombre"]
            return redirect(url_for("home"))
        else:
            return "Error: Contraseña incorrecta. <a href='/login'>Reintentar</a>"
    else:
        return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/recover")
def recover_redirect():
    return redirect(url_for("forgot_password"))

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        usuario_input = request.form.get("usuario").strip()
        user = users_collection.find_one({
            "$or": [
                {"email": {"$regex": f"^{usuario_input}$", "$options": "i"}},
                {"nombre": {"$regex": f"^{usuario_input}$", "$options": "i"}}
            ]
        })
        if not user:
            return "No se encontró ningún usuario con ese nombre o email. <a href='/forgot-password'>Volver</a>"
        token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(minutes=30)
        resets_collection.insert_one({
            "user_id": user["_id"],
            "token": token,
            "expires_at": expires_at,
            "used": False
        })
        reset_link = url_for("reset_password", token=token, _external=True)
        msg = Message("Recuperación de contraseña", recipients=[user["email"]])
        msg.body = (f"Hola {user['nombre']},\n\nPara restablecer tu contraseña, haz clic en el siguiente enlace:\n"
                    f"{reset_link}\n\nEste enlace caduca en 30 minutos.")
        mail.send(msg)
        return "Se ha enviado un enlace de recuperación a tu email. <a href='/login'>Inicia Sesión</a>"
    else:
        return render_template("forgot_password.html")

@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    token = request.args.get("token") or request.form.get("token")
    if not token:
        return "Token no proporcionado."
    reset_info = resets_collection.find_one({"token": token})
    if not reset_info:
        return "Token inválido o inexistente."
    if reset_info["used"]:
        return "Este token ya ha sido utilizado."
    if datetime.utcnow() > reset_info["expires_at"]:
        return "Token caducado."
    if request.method == "POST":
        new_pass = request.form.get("password").strip()
        hashed = generate_password_hash(new_pass)
        user_id = reset_info["user_id"]
        users_collection.update_one({"_id": user_id}, {"$set": {"password": hashed}})
        resets_collection.update_one({"_id": reset_info["_id"]}, {"$set": {"used": True}})
        return "Contraseña actualizada con éxito. <a href='/login'>Inicia Sesión</a>"
    else:
        return render_template("reset_password_form.html", token=token)

# -------------------------------------------
# RUTAS PARA GESTIÓN DE TABLAS (SpreadSheets)
# -------------------------------------------
@app.route("/")
def home():
    if "usuario" not in session:
        return render_template("welcome.html")  # Mostrar página de bienvenida si no hay sesión iniciada

    if "selected_table" in session:
        return redirect(url_for("catalog"))
    else:
        return redirect(url_for("tables"))

@app.route("/tables", methods=["GET", "POST"])
def tables():
    if "usuario" not in session:
        return redirect(url_for("login"))
    owner = session["usuario"]
    if request.method == "POST":
        table_name = request.form.get("table_name", "").strip()
        headers_str = request.form.get("table_headers", "").strip()
        headers = [h.strip() for h in headers_str.split(",")] if headers_str else ["Número", "Descripción", "Peso", "Valor", "Imagen1", "Imagen2", "Imagen3"]
        filename = f"table_{secrets.token_hex(8)}.xlsx"
        filepath = os.path.join(SPREADSHEET_FOLDER, filename)
        wb = Workbook()
        hoja = wb.active
        hoja.append(headers)
        wb.save(filepath)
        spreadsheets_collection.insert_one({"owner": owner, "name": table_name, "filename": filename, "headers": headers, "created_at": datetime.utcnow()})
        return redirect(url_for("tables"))
    tables = list(spreadsheets_collection.find({"owner": session["usuario"]}))
    return render_template("tables.html", tables=tables)

@app.route("/select_table/<table_id>")
def select_table(table_id):
    if "usuario" not in session:
        return redirect(url_for("login"))

    table = spreadsheets_collection.find_one({"_id": ObjectId(table_id)})

    if not table:
        flash("Tabla no encontrada.", "error")
        return redirect(url_for("tables"))

    session["selected_table"] = table["filename"]

    # Si la tabla tiene encabezados personalizados, usarlos. Si no, usar un valor por defecto.
    session["selected_headers"] = table.get("headers", ["Número", "Descripción", "Peso", "Valor", "Imagen1", "Imagen2", "Imagen3"])

    return redirect(url_for("catalog"))

# -------------------------------------------
# RUTAS DEL CATÁLOGO (Excel e imágenes) para la tabla seleccionada
# -------------------------------------------

@app.route("/catalog", methods=["GET", "POST"])
def catalog():
    if "usuario" not in session:
        return redirect(url_for("login"))

    spreadsheet_path = get_current_spreadsheet()
    if not spreadsheet_path or not os.path.exists(spreadsheet_path):
        return redirect(url_for("tables"))

    if request.method == "POST":
        form_data = {k.strip(): v.strip() for k, v in request.form.items()}
        print(f"Datos recibidos en el formulario: {form_data}")

        headers = session.get("selected_headers", ["Número", "Descripción", "Peso", "Valor", "Imagen1", "Imagen2", "Imagen3"])

        if "Número" not in form_data or not form_data["Número"]:
            return render_template("index.html", error_message="Error: Sin número.")

        data = leer_datos_excel(spreadsheet_path)

        # Verificar si el número ya existe
        if any(item["Número"] == form_data["Número"] for item in data):
            return render_template("index.html", data=data, error_message="Error: Ese número ya existe.")

        # Crear el nuevo registro
        nuevo_registro = {header: form_data.get(header, "") for header in headers}

        # Manejo de imágenes
        files = request.files.getlist("imagenes")
        rutas_imagenes = [None, None, None]
        for i, file in enumerate(files[:3]):
            if file and allowed_file(file.filename):
                fname = secure_filename(file.filename)
                fpath = os.path.join(app.config["UPLOAD_FOLDER"], fname)
                file.save(fpath)
                rutas_imagenes[i] = os.path.join("imagenes_subidas", fname)

        nuevo_registro["Imagenes"] = rutas_imagenes

        data.append(nuevo_registro)
        escribir_datos_excel(data, spreadsheet_path)

        return render_template("index.html", data=data, success_message="Registro agregado con éxito.")

    else:
        data = leer_datos_excel(spreadsheet_path)
        return render_template("index.html", data=data)

@app.route("/editar/<numero>", methods=["GET", "POST"])
def editar(numero):
    if "usuario" not in session:
        return redirect(url_for("login"))

    spreadsheet_path = get_current_spreadsheet()
    if not spreadsheet_path or not os.path.exists(spreadsheet_path):
        return redirect(url_for("tables"))

    # Obtener encabezados de la sesión
    headers = session.get("selected_headers", [])
    
    # Leer datos actuales del archivo
    data = leer_datos_excel(spreadsheet_path)

    # Encontrar el registro correspondiente
    idx_encontrado = None
    registro = None
    for i, item in enumerate(data):
        if item["numero"] == str(numero):
            idx_encontrado = i
            registro = item
            break

    if request.method == "GET":
        if not registro:
            return f"No existe el número {numero}."
        return render_template("editar.html", registro=registro, headers=headers)

    if request.method == "POST":
        if idx_encontrado is None:
            return f"No existe el número {numero}."

        delete_flag = request.form.get("delete_record")
        if delete_flag == "on":
            data.pop(idx_encontrado)
            # Reasignar los números
            for i, row in enumerate(data, start=1):
                row["numero"] = str(i)

            escribir_datos_excel(data, spreadsheet_path)
            return redirect(url_for("catalog"))

        # Crear nuevo diccionario basado en encabezados personalizados
        nuevo_registro = {header: request.form.get(header, "").strip() for header in headers}

        # Manejo de imágenes
        files = request.files.getlist("imagenes")
        rutas_imagenes = registro.get("imagenes", [None] * 3)

        for i, file in enumerate(files[:3]):
            if file and allowed_file(file.filename):
                fname = secure_filename(file.filename)
                fpath = os.path.join(app.config["UPLOAD_FOLDER"], fname)
                file.save(fpath)
                rutas_imagenes[i] = os.path.join("imagenes_subidas", fname)

        # Eliminar imágenes si se ha marcado la opción
        for i in range(3):
            if request.form.get(f"remove_img{i+1}") == "on":
                rutas_imagenes[i] = None

        # Asignar valores actualizados al registro
        nuevo_registro["imagenes"] = rutas_imagenes

        # Reemplazar el registro en la lista de datos
        data[idx_encontrado] = nuevo_registro

        # Guardar de nuevo en Excel
        escribir_datos_excel(data, spreadsheet_path)

        return redirect(url_for("catalog"))

@app.route("/delete_table/<table_id>", methods=["POST"])
def delete_table(table_id):
    if "usuario" not in session:
        return redirect(url_for("login"))
    # Buscamos la tabla en la colección, asegurándonos que el usuario actual es el propietario.
    table = spreadsheets_collection.find_one({"_id": ObjectId(table_id), "owner": session["usuario"]})
    if not table:
        flash("Tabla no encontrada o no tienes permiso para eliminarla.", "error")
        return redirect(url_for("tables"))
    
    # Construir la ruta absoluta del archivo Excel
    file_path = os.path.join(SPREADSHEET_FOLDER, table["filename"])
    # Si el archivo existe, lo eliminamos del sistema de archivos.
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            flash(f"Error al eliminar el archivo: {e}", "error")
            return redirect(url_for("tables"))
    
    # Eliminamos el documento de la colección en MongoDB
    spreadsheets_collection.delete_one({"_id": ObjectId(table_id)})
    
    # Si la tabla eliminada era la seleccionada en sesión, la removemos de la sesión.
    if session.get("selected_table") == table["filename"]:
        session.pop("selected_table", None)
    
    flash("Tabla eliminada exitosamente.", "success")
    return redirect(url_for("tables"))

@app.route("/descargar-excel")
def descargar_excel():
    if "usuario" not in session:
        return redirect(url_for("login"))
    spreadsheet_path = get_current_spreadsheet()
    if not spreadsheet_path or not os.path.exists(spreadsheet_path):
        return "El Excel no existe aún."
    temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    with zipfile.ZipFile(temp_zip.name, "w") as zf:
        zf.write(spreadsheet_path, arcname=os.path.basename(spreadsheet_path))
        data = leer_datos_excel(spreadsheet_path)
        image_paths = set()
        for row in data:
            for ruta in row["imagenes"]:
                if ruta:
                    absolute_path = os.path.join(app.root_path, ruta)
                    if os.path.exists(absolute_path):
                        image_paths.add(absolute_path)
        for img_path in image_paths:
            arcname = os.path.join("imagenes", os.path.basename(img_path))
            zf.write(img_path, arcname=arcname)
    return send_from_directory(directory=os.path.dirname(temp_zip.name),
                               path=os.path.basename(temp_zip.name),
                               as_attachment=True,
                               download_name="catalogo.zip")

@app.route("/imagenes_subidas/<path:filename>")
def uploaded_images(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

# -------------------------------------------
# MAIN
# -------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)